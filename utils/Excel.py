# coding:utf-8

import openpyxl

import sys

sys.path.append("..")

class ExcelUtils(object):
    def __init__(self,path="..\\testData\\importOrderList.xls"):
        self.path =path

    def read_excel(self,rowOrCol="row",sheetName="sheet1"):
        data = openpyxl.load_workbook(self.path)  # 打开电影.xlsx文件读取数据
        table = data.get_sheet_by_name(sheetName)  # 读取sheet表单
        # 或者通过表单名称获取 table = data.sheet_by_name(u'Sheet1')
        print("输出表格行数=",table.nrows)  # 输出表格行数
        print("输出表格列数=",table.ncols)  # 输出表格列数
        if rowOrCol=="row":
            print(table.row_values(0))  # 输出第一行
        else:
            print(table.col_values(0))  # 输出第一列
        # print(table.cell(0, 2).value)  # 输出元素（0,2）的值

    def order_number_addOne(self):
        colName="订单号"
        addOne=int(1)
        wb = openpyxl.load_workbook(self.path)  # 打开电影.xlsx文件读取数据
        sheet_names = wb.sheetnames
        sheet1 = wb[sheet_names[0]]  # 读取sheet表单
        cols=sheet1.max_column  #列
        # print(cols)
        rows=sheet1.max_row  #行
        # print(rows)
        for row in range(2,rows+1):
            print(sheet1.cell(row=row, column=3).value)
            order_number=sheet1.cell(row=row, column=3).value
            new_order_number=int(order_number)+addOne
            print(new_order_number)
            sheet1.cell(row=row, column=3,value=str(new_order_number))
            wb.save(path)

if __name__ == '__main__':
    path = "..\\testData\\importOrderList.xlsx"
    # path = "..\\testData\\test.xls"
    testExcel = ExcelUtils(path)
    testExcel.order_number_addOne()
    # testExcel.read_excel()
